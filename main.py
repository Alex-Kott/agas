# coding=utf-8
import re
import sys
from datetime import datetime, timedelta

sys.path.insert(0, './venv/lib/python2.7/site-packages/')

from openpyxl import load_workbook
from pathlib import Path
from bs4 import BeautifulSoup
from openpyxl.styles import PatternFill


def is_interesting(tr, azs_prefixes):
    for prefix in azs_prefixes:
        if tr.text.find(prefix) != -1:
            return True
    return False


def get_agas_date_range(trs):
    for tr in trs:
        line_text = tr.text.encode('utf-8')
        if line_text.find("Медиа") != -1:
            result = re.search(r'(?<=_)((\d{2}.\d{2})(.\d{2})?)\s-\s(\d{2}.\d{2}.\d{4})', line_text)
            if result is not None:
                end_date = datetime.strptime(result.group(4), "%d.%m.%Y")
                if result.group(1) == result.group(2):
                    start_date = datetime.strptime(result.group(1), "%d.%m")
                    start_date = start_date.replace(year=int(end_date.strftime("%Y")))
                else:
                    start_date = datetime.strptime(result.group(1), "%d.%m.%Y")

                return datetime.date(start_date), datetime.date(end_date)


def get_day_difference(text):
    result = re.search(r'(?<=\()\d{1,}(?=\s\D+)', text)

    return int(result.group(0))


def get_order_date_range(text):
    result = re.search(r'(?<=\d{2}-)\d{2}.\d{2}', text)
    end_date = datetime.strptime(result.group(0), "%d.%m")
    end_date = end_date.replace(year=int(agas_start_date.strftime("%Y")))
    day_difference = get_day_difference(text)
    start_date = end_date - timedelta(day_difference - 1)  # особенности подсчёта в отчёте

    return datetime.date(start_date), datetime.date(end_date)


def parse_report(file_name):
    global agas_start_date, agas_end_date
    azs_prefixes = ['ALK', 'CHO', 'EKA', 'HMO', 'IVO', 'KDK', 'KEO', 'KYK', 'MSK', 'NNO', 'NSO',
                    'OMO', 'SPB', 'SVO', 'TUO', 'YAR', 'IAR', 'KRR', 'MOW', 'MSK', 'NN', 'RYZ',
                    'CEK', 'MSK', 'NN', 'SPB', 'TUO', 'YAR']
    with open(str(file_name)) as file:
        soup = BeautifulSoup(file.read(), "lxml")

    tbody = soup.find("tbody")
    trs = tbody.find_all("tr")
    interesting_tr = []

    # agas_start_date, agas_end_date = get_agas_date_range(trs)

    for tr in trs:
        if is_interesting(tr, azs_prefixes):
            interesting_tr.append(tr)

    data = {}
    for entry in interesting_tr:
        tds = entry.find_all("td")
        player_name = tds[1].text.strip('\n')
        try:
            azs_number = int(re.findall(r'(?<=\D{3}[_-])\d+', player_name)[0])
            data[azs_number] = int(tds[5].text)
        except:
            pass

    return data


def is_empty(s, n):
    print ws['{}{}'.format(s, n)]
    if ws['{}{}'.format(s, n)] == '':
        return True

    return True


def parse_arguments(arguments):
    if len(arguments) != 3:
        raise Exception("Must be 2 argument (.html-file and .xlsx-file)")

    args = [Path(arg) for arg in arguments]

    for arg in args:
        if arg.suffix == ".html":
            agas_file_name = arg
        elif arg.suffix == ".xlsx":
            order_file_name = arg
        elif arg.suffix == ".py":
            pass
        else:
            raise Exception("Unknown file")

    return agas_file_name, order_file_name


def print_row(row):
    # for cell in row:
    #     print type(cell.value)
    try:
        line = ['({}) {}'.format(cell.coordinate, cell.value.encode('utf-8'))
                for cell in row if cell.value is not None]
        print "    ".join(line)
    except:
        pass


# def suitable_date_range(row):
#     if order_start_date == agas_start_date and order_end_date == agas_end_date:
#         return True
#     return False


if __name__ == "__main__":
    errors = []
    (agas_file_name, order_file_name) = parse_arguments(sys.argv)

    report = parse_report(agas_file_name)
    print(report[14205])
    wb = load_workbook(str(order_file_name))
    ws = wb.active

    flag = False
    for row in ws.iter_rows():
        if row[1].value == u"№ АЗС:":
            flag = True
            continue
            # order_start_date, order_end_date = get_order_date_range(row[12].value)
            # if suitable_date_range(row):
            #     flag = True
            #     continue
            # else:
            #     errors.append("Даты отчёта ({} - {}) не соответствуют целевому диапазону дат.".format(
            #         agas_start_date, agas_end_date
            #     ))
        elif row[1].value == None:
            flag = False
        if flag == False:
            continue

        azs_code = row[1].value

        # if type(row[12].value) is long:
        #     n = row[12].value
        #
        # elif type(row[12].value) is str:
        #     try:
        #         n = eval(row[12].value.strip('='))
        #     except:
        #         pass

        row[15].value = report[azs_code]
        # row16 = report[azs_code] - n
        # if row16 < 0:
        #     row[16].fill = PatternFill(fill_type='solid',
        #                                start_color='FF9900',
        #                                end_color='FF9900')



    wb.save(str(order_file_name))
